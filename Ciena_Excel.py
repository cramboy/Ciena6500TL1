# Ciena_Excel.py
'''
Module used to read from and write to Excel documents
'''
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pprint  # CLI feedback troubleshooting
from prettytable import PrettyTable
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#


def getTabs(ExcelFileName):
    '''Returns a list of the tab names ExcelFileName'''
    # ExcelFileName is an Excel Workbook that holds information about the Ciena 
    # 6500 DWDM devices in a provider network. If provider networks contain 
    # multiple DWDM networks (DWDM topologies that are logically unrelated), use 
    # separate tabs to group nodes together. Be aware that grouped nodes must 
    # use the same credentials for login.

    # Ciena_Excel.py uses specific naming and formatting conventions that must be 
    # adhered to for the programming in this module to work correctly.

    # Workbook: must be an 'xlsx' file
    # Tab names: must be one word used to describe a specific DWDM network  
    #    Examples: WestRing, ABCDWDM, Z006NET
    # Tab names: must begin with an alpha character
    # Worksheets: Row 1 is used as column headers only
    # Worksheets: Column A is empty
    # Worksheets: Column B holds the node names
    # Worksheets: Column C is the management IPv4 of the node

    excelWorkbook = openpyxl.load_workbook(ExcelFileName)
    tabList = excelWorkbook.sheetnames
    return(tabList)


def getIPadd(ExcelFileName):
    '''Returns a list of the IP addresses in each tab of the file.xlsx'''
    excelWorkbook = openpyxl.load_workbook(ExcelFileName)
    # names of all the tabs in the workbook
    tabList = excelWorkbook.sheetnames

    workBookIPaddresses = []  # All ip addresses in the workbook
    # Get the IP addresses in each tab
    for tabIndex, tabName in enumerate(tabList):
        worksheet = excelWorkbook[tabName]  # a particular tab
        tabIPaddresses = []  # The IP addresses in that tab
        # Top row is used for column header labels, start on the second row
        # max_row is the last row that has data in a cell
        for eachRow in range(2, worksheet.max_row + 1):
            if worksheet["C" + str(eachRow)].value != None:
                # IP addresses are in column C
                tabIPaddresses.append(worksheet["C" + str(eachRow)].value)
        workBookIPaddresses.append(tabIPaddresses)
    return(workBookIPaddresses)


def getNodeNames(ExcelFileName):
    excelWorkbook = openpyxl.load_workbook(ExcelFileName)
    # names of all the tabs in the workbook
    tabList = excelWorkbook.sheetnames

    workBookNodeNames = []  # all tabIPaddresses for each tab in the workbook

    for tabIndex, tabName in enumerate(tabList):
        worksheet = excelWorkbook[tabName]  # a particular tab
        tabNodeNames = []  # the node names in that tab

        # Top row is used for column header labels, start on the second row
        # max_row is the last row that has data in a cell
        for eachRow in range(2, worksheet.max_row + 1):
            if worksheet["B" + str(eachRow)].value != None:
                # IP addresses are in column C
                tabNodeNames.append(worksheet["B" + str(eachRow)].value)
        workBookNodeNames.append(tabNodeNames)
    return(workBookNodeNames)

# -------------------------------------------------------------------------------------------------------


def chooseTab(ExcelFileName):
    '''Prints a selection menu to commandline allowing the user to select 
       the network they want to login to
    '''

    # alphabet = list of letters in the alphabet for the user to make a selection
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    excelWorkbook = openpyxl.load_workbook(ExcelFileName)
    # names of all the tabs in the workbook
    workBookTabNames = excelWorkbook.sheetnames

    workBookIPaddresses = []  # All ip addresses in the workbook
    # Get the IP addresses in each tab
    for tabIndex, tabName in enumerate(workBookTabNames):
        worksheet = excelWorkbook[tabName]  # a particular tab
        tabIPaddresses = []  # The IP addresses in that tab
        # Top row is used for column header labels, start on the second row
        # max_row is the last row that has data in a cell
        for eachRow in range(2, worksheet.max_row + 1):
            if worksheet["C" + str(eachRow)].value != None:
                # IP addresses are in column C
                tabIPaddresses.append(worksheet["C" + str(eachRow)].value)
        workBookIPaddresses.append(tabIPaddresses)

    workBookNodeNames = []  # All node names in the workbook
    for tabIndex, tabName in enumerate(workBookTabNames):
        worksheet = excelWorkbook[tabName]  # a particular tab
        tabNodeNames = []  # the node names in that tab

        for eachRow in range(2, worksheet.max_row + 1):
            if worksheet["B" + str(eachRow)].value != None:
                # IP addresses are in column C
                tabNodeNames.append(worksheet["B" + str(eachRow)].value)
        workBookNodeNames.append(tabNodeNames)

    loop = True
    print()
    print("SELECT NETWORK")
    while loop:
        # Prompt the user to make a tab/network selection
        print("Type the letter shown in brackets to make your selection")
        print("Type 'ESC' to abort")
        selectionAlphabet = []  # used later to hold the letters used to make this
        # particular menu

        tableData = []
        # For each tab
        for tabIndex, eachTab in enumerate(workBookTabNames):

            # print the selection t the screen
            print(f"   [{alphabet[tabIndex]}] {eachTab}")
            tableData.append([alphabet[tabIndex], eachTab])  # [A, WDM]

            # Create a short list of the letters used for selections
            # This is used below to see if the user typed a letter that was
            # actually one of the choices
            selectionAlphabet.append(alphabet[tabIndex])

        # User makes a selection
        tabSelection = input("\nSELECT> ")  # user types the letter
        tabSelection = tabSelection.upper()

        # ------------
        # Add code here to verify that a valid selection was made
        # ------------

        netNodeIPs = []
        if tabSelection in selectionAlphabet:
            alphaIndex = selectionAlphabet.index(tabSelection)

            networkName = workBookTabNames[alphaIndex]
            networkNodes = workBookNodeNames[alphaIndex]
            networkIPs = workBookIPaddresses[alphaIndex]

            print()
            print("CONFIRM SELECTION")
            print(f"{networkName} network selected")
            print("Is this selection correct? Type Yes to proceed or No to reselect")
            print()
            confirm = input("CONFIRM> ")
            confirm = confirm.upper()

            if confirm[0] == "Y":
                netNodeIPs.append(networkName)
                netNodeIPs.append(networkNodes)
                netNodeIPs.append(networkIPs)
                loop = False
                return(netNodeIPs)
            else:
                loop = True
